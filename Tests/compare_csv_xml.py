import pandas as pd
import csv
from openpyxl import Workbook
from datetime import datetime
from xml.etree import ElementTree as ET
import re
from S3_handler.S3_reader import read_xml_from_s3  # Import function to read XML from S3
from CSV_handler.CSV_reader import read_csv_from_local  # Import function to read JSON from local

# Define test parameters
S3_BUCKET = "testlocxml"
XML_FILE_KEY = "order2_with_dates.xml"
#LOCAL_CSV_FILE_PATH_HEADER = "scr/order1.csv"
#LOCAL_CSV_FILE_PATH_LINE = "scr/orderline1.csv"

def compare_csv_xml():
    
    # Read XML from S3
    xml_root = read_xml_from_s3(S3_BUCKET, XML_FILE_KEY)
   
    wb=Workbook()
    ws=wb.active

    # Write Order fields
    ws["A1"] = "System"
    ws["A2"] = "OrderID"
    ws["A3"] = "Status"
    ws["A4"] = "CustomerID"
    ws["A5"] = "CustomerName"
    ws["A6"] = "Email"
    ws["A7"] = "Phone"
    ws["A8"] = "OrderDate"
    ws["A9"] = "TotalAmount"
    ws["A10"] = "Currency"
    ws["A11"] = "ShippingMethod"
    ws["A12"] = "TrackingNumber"
    ws["A13"] = "Street"
    ws["A14"] = "City"
    ws["A15"] = "State"
    ws["A16"] = "Zip"
    ws["A17"] = "Country"
    ws["A18"] = "LineNumber"
    ws["A19"] = "ProductID"
    ws["A20"] = "ProductName"
    ws["A21"] = "Category"
    ws["A22"] = "Quantity"
    ws["A23"] = "UnitPrice"
    ws["A24"] = "Currency"
    ws["A25"] = "DiscountAmount"
    ws["A26"] = "DiscountPercentage"
    
    
    order_checked = input("Enter the order number to check (e.g., 2002): ")
    order_line_checked = input("Enter the order line number to check (e.g., 1): ")
    
    
    order = xml_root.find(f"./Order[@id='{order_checked}']")
    if order is None:
        print(f"Order with ID {order_checked} not found.")
        return

    order_line = order.find(f"./OrderLines/OrderLine[@lineNumber='{order_line_checked}']")
    if order_line is None:
        print(f"Order line number {order_line_checked} not found for order {order_checked}.")
        return
    
    customer = order.find("Customer")
    order_details = order.find("OrderDetails")
    shipping = order.find("Shipping")
    address = shipping.find("Address")
    product = order_line.find("Product")
 
    # XML fields

    ws["B1"] = "XML File Fields"
    ws["B2"] = order.get("id")
    ws["B3"] = order.get("status")
    ws["B4"] = customer.find("CustomerID").text
    ws["B5"] = customer.find("Name").text
    ws["B6"] = customer.find("Email").text
    ws["B7"] = customer.find("Phone").text
    ws["B8"] = order_details.find("OrderDate").text
    ws["B9"] = order_details.find("TotalAmount").text
    ws["B10"] = order_details.find("TotalAmount").get("currency")
    ws["B11"] = shipping.find("Method").text
    ws["B12"] = shipping.find("TrackingNumber").text
    ws["B13"] = address.find("Street").text
    ws["B14"] = address.find("City").text
    ws["B15"] = address.find("State").text
    ws["B16"] = address.find("Zip").text
    ws["B17"] = address.find("Country").text
    ws["B18"] = order_line.get("lineNumber")
    ws["B19"] = product.find("ProductID").text
    ws["B20"] = product.find("Name").text
    ws["B21"] = product.find("Category").text
    ws["B22"] = order_line.find("Quantity").text
    ws["B23"] = order_line.find("UnitPrice").text
    ws["B24"] = order_line.find("UnitPrice").get("currency")
    ws["B25"] = order_line.find("Discount").text
    ws["B26"] = order_line.find("Discount").get("percentage")
    
    LOCAL_CSV_FILE_PATH_HEADER = "C:\\Users\\Documents\\Files_for_data_comparison\\order_header_fields.csv"
    LOCAL_CSV_FILE_PATH_LINE = "C:\\Users\\Documents\\Files_for_data_comparison\\order_line_fields.csv"
    orders_df, order_lines_df = read_csv_from_local (LOCAL_CSV_FILE_PATH_HEADER,LOCAL_CSV_FILE_PATH_LINE)
        
    filtered_order = orders_df[orders_df["OrderID"] == int(order_checked)]
    if filtered_order.empty:
        print(f"No matching order found in CSV for OrderID: {order_checked}")
        return
    order_csv = filtered_order.iloc[0]

    filtered_line = order_lines_df[
        (order_lines_df["OrderID"] == int(order_checked)) &
        (order_lines_df["LineNumber"] == int(order_line_checked))
    ]
    if filtered_line.empty:
        print(f"No matching order line for OrderID: {order_checked}, LineNumber: {order_line_checked}")
        return
    order_line_csv = filtered_line.iloc[0]

    # Column C - CSV fields
    ws["C1"] = "CSV File Fields"
    ws["C2"] = order_csv["OrderID"]
    ws["C3"] = order_csv["Status"]
    ws["C4"] = order_csv["CustomerID"]
    ws["C5"] = order_csv["CustomerName"]
    ws["C6"] = order_csv["Email"]
    ws["C7"] = order_csv["Phone"]
    ws["C8"] = order_csv["OrderDate"]
    ws["C9"] = order_csv["TotalAmount"]
    ws["C10"] = order_csv["Currency"]
    ws["C11"] = order_csv["ShippingMethod"]
    ws["C12"] = order_csv["TrackingNumber"]
    ws["C13"] = order_csv["Street"]
    ws["C14"] = order_csv["City"]
    ws["C15"] = order_csv["State"]
    ws["C16"] = order_csv["Zip"]
    ws["C17"] = order_csv["Country"]
    ws["C18"] = order_line_csv["LineNumber"]
    ws["C19"] = order_line_csv["ProductID"]
    ws["C20"] = order_line_csv["ProductName"]
    ws["C21"] = order_line_csv["Category"]
    ws["C22"] = order_line_csv["Quantity"]
    ws["C23"] = order_line_csv["UnitPrice"]
    ws["C24"] = order_line_csv["Currency"]
    ws["C25"] = order_line_csv["DiscountAmount"]
    ws["C26"] = order_line_csv["DiscountPercentage"]

    wb.save("./Tests/xml_values_line_by_line.xlsx")
    
     # Column D - XML fields Transformed
    ws["D1"] = "XML File Fields Transformed"
    ws["D2"] = int(ws["B2"].value)
    ws["D3"] = ws["B3"].value.upper() # Convert status to uppercase
    ws["D4"] = int(ws["B4"].value)
    ws["D5"] = ws["B5"].value
    ws["D6"] = ws["B6"].value
    ws["D7"] = int(re.sub(r"\D", "", ws["B7"].value )) # Remove non-numeric characters from phone
    date_obj = datetime.strptime(ws["B8"].value, "%Y-%m-%d")
    ws["D8"] = date_obj.strftime("%d-%m-%Y")
    ws["D9"] = round(float(ws["B9"].value))
    ws["D10"] = ws["B10"].value
    ws["D11"] = ws["B11"].value
    ws["D12"] = ws["B12"].value
    ws["D13"] = ws["B13"].value
    ws["D14"] = ws["B14"].value
    ws["D15"] = ws["B15"].value
    ws["D16"] = int(ws["B16"].value)
    ws["D17"] = ws["B17"].value
    ws["D18"] = int(ws["B18"].value)
    ws["D19"] = int(ws["B19"].value)
    ws["D20"] = ws["B20"].value
    ws["D21"] = ws["B21"].value
    ws["D22"] = ws["B22"].value
    ws["D23"] = round(float(ws["B23"].value))
    ws["D24"] = ws["B24"].value
    ws["D25"] = ws["B25"].value
    ws["D26"] = ws["B26"].value
    
       # cpmpare XML and CSV fields
    for row in range(2, 27):
        
        csv_value = ws[f"C{row}"].value
        transformed_value = ws[f"D{row}"].value
        
        if transformed_value == csv_value:
            ws[f"E{row}"] = "Match"
        else:
            ws[f"E{row}"] = "Mismatch"
        

    wb.save("./Tests/xml_values_line_by_line.xlsx")
           
    
if __name__ == "__main__":
    compare_csv_xml()
# This code compares the fields from an XML file stored in S3 with those from a CSV file stored locally.
# It reads the XML and CSV files, extracts relevant fields, and writes them into an Excel








   

    
   
